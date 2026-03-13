"""
Script to create a placeholder template.xlsx file for MySmartReport.
Run this once to generate the template.
Replace with the real template when available.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def create_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "RPP Mingguan"

    # --- Styling ---
    header_font = Font(name="Arial", size=12, bold=True)
    label_font = Font(name="Arial", size=10, bold=True)
    normal_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_text = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    wrap_align = Alignment(wrap_text=True, vertical="top")

    # --- Column widths ---
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 20

    # --- Row 1: Title ---
    ws.merge_cells("A1:F1")
    ws["A1"] = "RANCANGAN PENGAJARAN & PEMBELAJARAN MINGGUAN"
    ws["A1"].font = header_text
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # --- Row 2: blank spacer ---
    ws.row_dimensions[2].height = 8

    # --- Row 3: Minggu / Kumpulan ---
    ws["A3"] = "Minggu:"
    ws["A3"].font = label_font
    ws.merge_cells("A3:B3")
    ws["C3"].font = normal_font
    ws["C3"].border = thin_border

    ws["E3"] = "Kumpulan:"
    ws["E3"].font = label_font
    ws["F3"].font = normal_font
    ws["F3"].border = thin_border

    # --- Row 4: Tarikh / Kehadiran ---
    ws["A4"] = "Tarikh:"
    ws["A4"].font = label_font
    ws.merge_cells("A4:B4")
    ws["C4"].font = normal_font
    ws["C4"].border = thin_border

    ws["E4"] = "Kehadiran:"
    ws["E4"].font = label_font
    ws["F4"].font = normal_font
    ws["F4"].border = thin_border

    # --- Row 5: Masa ---
    ws["A5"] = "Masa:"
    ws["A5"].font = label_font
    ws.merge_cells("A5:B5")
    ws["C5"].font = normal_font
    ws["C5"].border = thin_border

    # --- Row 6: blank spacer ---
    ws.row_dimensions[6].height = 8

    # --- Row 7: Topik ---
    ws["A7"] = "Topik:"
    ws["A7"].font = label_font
    ws.merge_cells("B7:F7")
    ws["B7"].font = normal_font
    ws["B7"].border = thin_border
    ws["B7"].alignment = wrap_align
    ws.row_dimensions[7].height = 40

    # --- Row 8: blank spacer ---
    ws.row_dimensions[8].height = 8

    # --- Row 9: HPK ---
    ws["A9"] = "HPK:"
    ws["A9"].font = label_font
    ws.merge_cells("B9:F9")
    ws["B9"].font = normal_font
    ws["B9"].border = thin_border
    ws["B9"].alignment = wrap_align
    ws.row_dimensions[9].height = 30

    # --- Row 10: blank spacer ---
    ws.row_dimensions[10].height = 8

    # --- Row 11-13: Hasil Pembelajaran Topik ---
    section_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    ws.merge_cells("A11:F11")
    ws["A11"] = "Hasil Pembelajaran Topik"
    ws["A11"].font = label_font
    ws["A11"].fill = section_fill
    ws["A11"].border = thin_border

    ws.merge_cells("B11:F13")
    ws["B11"].font = normal_font
    ws["B11"].border = thin_border
    ws["B11"].alignment = wrap_align
    ws.row_dimensions[11].height = 25
    ws.row_dimensions[12].height = 25
    ws.row_dimensions[13].height = 25

    # --- Row 14-17: Strategi / Aktiviti P&P ---
    ws.merge_cells("A14:F14")
    ws["A14"] = "Strategi / Aktiviti Pengajaran & Pembelajaran"
    ws["A14"].font = label_font
    ws["A14"].fill = section_fill
    ws["A14"].border = thin_border

    ws.merge_cells("B14:F17")
    ws["B14"].font = normal_font
    ws["B14"].border = thin_border
    ws["B14"].alignment = wrap_align
    ws.row_dimensions[14].height = 25
    ws.row_dimensions[15].height = 25
    ws.row_dimensions[16].height = 25
    ws.row_dimensions[17].height = 25

    # --- Row 18-20: Refleksi ---
    ws.merge_cells("A18:F18")
    ws["A18"] = "Refleksi"
    ws["A18"].font = label_font
    ws["A18"].fill = section_fill
    ws["A18"].border = thin_border

    ws.merge_cells("B18:F20")
    ws["B18"].font = normal_font
    ws["B18"].border = thin_border
    ws["B18"].alignment = wrap_align
    ws.row_dimensions[18].height = 25
    ws.row_dimensions[19].height = 25
    ws.row_dimensions[20].height = 25

    # Apply borders to all cells in the content area
    for row in ws.iter_rows(min_row=3, max_row=20, max_col=6):
        for cell in row:
            if cell.border == Border():
                cell.border = thin_border

    wb.save("templates/template.xlsx")
    print("template.xlsx created successfully in templates/")


if __name__ == "__main__":
    create_template()
