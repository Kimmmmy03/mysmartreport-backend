"""
MySmartReport — Input XLSX/PDF Parser Service (v3 — Bulletproof)

Three-layer extraction strategy:
  Layer 1: Anchor-based scanning (handles standard cell layouts)
  Layer 2: Pattern-based scanning (handles merged/unusual layouts)
  Layer 3: LLM-assisted extraction (handles garbled PDFs + scanned images)

Critical fields: nama_kursus, kod_kursus, pensyarah, kumpulan_diajar, tarikh.
Never crashes — returns None/empty for fields that can't be confidently extracted.
"""

import re
import asyncio
from io import BytesIO
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
import pdfplumber

from models import WeekData, UploadMetadata
from services.gemini_service import extract_file_content, is_exception_week


# ═══════════════════════════════════════════════════════════════════════════
#  CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════

BULAN_MELAYU = {
    1: "Januari", 2: "Februari", 3: "Mac", 4: "April",
    5: "Mei", 6: "Jun", 7: "Julai", 8: "Ogos",
    9: "September", 10: "Oktober", 11: "November", 12: "Disember",
}
BULAN_NAMA_TO_NUM = {name.lower(): num for num, name in BULAN_MELAYU.items()}

# Metadata anchors: (anchor_keywords, field_name)
_METADATA_ANCHORS: list[tuple[list[str], str]] = [
    (["nama kursus", "mata pelajaran", "course name"], "nama_kursus"),
    (["kod kursus", "course code"], "kod_kursus"),
    (["nama pensyarah", "pensyarah", "lecturer"], "pensyarah"),
    (["kumpulan diajar", "kumpulan ajar"], "kumpulan_diajar"),
    (["program"], "program"),
    (["semester"], "semester"),
    (["tahun"], "tahun"),
    (["ambilan", "intake"], "ambilan"),
    (["jabatan", "unit"], "jabatan"),
    (["jumlah kredit", "kredit"], "jumlah_kredit"),
]

# Regex patterns for pattern-based detection (Layer 2)
_KOD_KURSUS_PATTERN = re.compile(
    r'\b([A-Z]{2,5}\d{3,5}[A-Z]?)\b'
)
# ALLCAPS name pattern (2+ words, each 2+ chars, with BIN/BINTI)
_NAMA_PATTERN = re.compile(
    r'\b([A-Z][A-Z .]+\s+(?:BIN(?:TI)?|B\.?|BT\.?)\s+[A-Z][A-Z .]+)\b'
)


# ═══════════════════════════════════════════════════════════════════════════
#  UTILITY FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════

def _clean(value) -> str:
    """Strip a value to a clean string, removing stray colons/whitespace/newlines."""
    if value is None:
        return ""
    s = str(value).strip()
    # Collapse internal newlines to spaces for metadata
    s = re.sub(r'[\r\n]+', ' ', s)
    # Remove leading/trailing colons, dashes, whitespace
    s = s.strip(" \t:-;")
    return s


def _cell_text(cell) -> str:
    """Get clean text from an openpyxl cell."""
    if cell.value is None:
        return ""
    return str(cell.value).strip()


def _matches_anchor(text: str, anchors: list[str]) -> bool:
    """Check if text fuzzy-matches any anchor keyword."""
    t = text.strip().lower().rstrip(":")
    for anchor in anchors:
        if anchor in t:
            return True
    return False


def _is_label_cell(text: str) -> bool:
    """Check if text looks like a label (contains a metadata anchor)."""
    t = text.strip().lower()
    for anchors, _ in _METADATA_ANCHORS:
        for a in anchors:
            if a in t:
                return True
    return False


def _split_kumpulan(value: str) -> list[str]:
    """Split a kumpulan string into a list of group names."""
    groups = re.split(r'[,;&/]', value)
    return [g.strip() for g in groups if g.strip()]


# ═══════════════════════════════════════════════════════════════════════════
#  DATE FORMATTING
# ═══════════════════════════════════════════════════════════════════════════

def _format_tarikh(tarikh: str) -> str:
    """
    Extract ONLY the START date from a tarikh string.
    Returns "d MMMM yyyy" in Bahasa Melayu.
    Returns empty string if parsing fails.
    """
    if not tarikh:
        return ""

    from datetime import datetime

    text = str(tarikh).strip()
    if not text:
        return ""

    # Normalize dashes
    text = text.replace("–", "-").replace("—", "-")

    # Pattern 1: "25-29 Ogos 2025" or "25 - 29 Ogos 2025"
    m = re.search(r'(\d{1,2})\s*-\s*\d{1,2}\s+([A-Za-zÀ-ÖØ-öø-ÿ]+)\s+(\d{4})', text)
    if m:
        day, month_raw, year = int(m.group(1)), m.group(2).strip(), int(m.group(3))
        month_name = BULAN_MELAYU.get(BULAN_NAMA_TO_NUM.get(month_raw.lower(), 0), month_raw)
        return f"{day} {month_name} {year}"

    # Pattern 2: "25 Ogos - 29 Ogos 2025"
    m = re.search(r'(\d{1,2})\s+([A-Za-zÀ-ÖØ-öø-ÿ]+)\s*-\s*\d{1,2}\s+[A-Za-zÀ-ÖØ-öø-ÿ]+\s+(\d{4})', text)
    if m:
        day, month_raw, year = int(m.group(1)), m.group(2).strip(), int(m.group(3))
        month_name = BULAN_MELAYU.get(BULAN_NAMA_TO_NUM.get(month_raw.lower(), 0), month_raw)
        return f"{day} {month_name} {year}"

    # Pattern 3: "25 Ogos 2025 - 29 Ogos 2025"
    m = re.search(
        r'(\d{1,2})\s+([A-Za-zÀ-ÖØ-öø-ÿ]+)\s+(\d{4})\s*-\s*\d{1,2}\s+[A-Za-zÀ-ÖØ-öø-ÿ]+\s+\d{4}',
        text,
    )
    if m:
        day, month_raw, year = int(m.group(1)), m.group(2).strip(), int(m.group(3))
        month_name = BULAN_MELAYU.get(BULAN_NAMA_TO_NUM.get(month_raw.lower(), 0), month_raw)
        return f"{day} {month_name} {year}"

    # Pattern 4: "25 Ogos 2025" (standalone)
    m = re.search(r'(\d{1,2})\s+([A-Za-zÀ-ÖØ-öø-ÿ]+)\s+(\d{4})', text)
    if m:
        day, month_raw, year = int(m.group(1)), m.group(2).strip(), int(m.group(3))
        month_num = BULAN_NAMA_TO_NUM.get(month_raw.lower(), 0)
        if month_num:
            return f"{day} {BULAN_MELAYU[month_num]} {year}"

    # Pattern 5: Numeric "25/08/2025" or "25-08-2025"
    m = re.search(r'(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})', text)
    if m:
        day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if year < 100:
            year += 2000
        if 1 <= month <= 12:
            return f"{day} {BULAN_MELAYU[month]} {year}"

    # Pattern 6: datetime object
    try:
        dt = datetime.fromisoformat(text.replace("Z", ""))
        return f"{dt.day} {BULAN_MELAYU[dt.month]} {dt.year}"
    except (ValueError, TypeError):
        pass

    return text


# ═══════════════════════════════════════════════════════════════════════════
#  LAYER 1: ANCHOR-BASED XLSX EXTRACTION
# ═══════════════════════════════════════════════════════════════════════════

def _extract_metadata_anchored(ws, max_row: int = 20) -> dict[str, str | list[str]]:
    """
    Scan worksheet for anchor words and extract adjacent values.
    Handles: label in col A → value in col C/D, or label:value in same cell.
    SKIPS cells that contain multiple anchor keywords (those are merged-label
    blocks that Layer 2 handles via line-offset mapping).
    """
    result: dict[str, str | list[str]] = {}

    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_row), values_only=False):
        for cell in row:
            ct = _cell_text(cell)
            if not ct:
                continue

            # Skip merged-label blocks (cells with 2+ anchor keywords)
            anchor_count = 0
            for anchors, _ in _METADATA_ANCHORS:
                if _matches_anchor(ct, anchors):
                    anchor_count += 1
            if anchor_count >= 2:
                continue  # Defer to Layer 2

            for anchors, field_name in _METADATA_ANCHORS:
                if field_name in result:
                    continue
                if not _matches_anchor(ct, anchors):
                    continue

                value = ""

                # Source A: After colon in same cell (e.g., "NAMA KURSUS: PSIKOLOGI")
                if ":" in ct:
                    _, _, after = ct.partition(":")
                    after = after.strip()
                    # Make sure it's not just another label fragment
                    if after and not _is_label_cell(after):
                        value = _clean(after)

                # Source B: Next non-empty cell to the right (skip other labels)
                if not value:
                    for c2 in row:
                        if c2.column <= cell.column:
                            continue
                        c2t = _cell_text(c2)
                        if not c2t:
                            continue
                        if _is_label_cell(c2t):
                            continue
                        value = _clean(c2t)
                        break

                # Source C: Cell directly below
                if not value:
                    try:
                        below = ws.cell(row=cell.row + 1, column=cell.column)
                        bt = _cell_text(below)
                        if bt and not _is_label_cell(bt):
                            value = _clean(bt)
                    except Exception:
                        pass

                if not value:
                    continue

                if field_name == "kumpulan_diajar":
                    result[field_name] = _split_kumpulan(value)
                else:
                    result[field_name] = value
                break

    return result


# ═══════════════════════════════════════════════════════════════════════════
#  LAYER 2: LABEL-COUNT OFFSET MAPPING FOR MERGED-LABEL LAYOUTS
# ═══════════════════════════════════════════════════════════════════════════

def _extract_metadata_patterns(ws, max_row: int = 20) -> dict[str, str | list[str]]:
    """
    For layouts where ALL labels are merged into one cell (e.g., Sample 3),
    the cell text contains multiple labels like:
        A2 = "PROGRAM: [1] SEMESTER:\\nNAMA PENSYARAH: JABATAN / UNIT:\\nNAMA KURSUS:\\nKOD KURSUS:"

    The values are in cells in an adjacent column, one per row, with each
    label getting the NEXT sequential row offset:
        D2 = PPISMP         (offset 0 → PROGRAM)
        D3 = I              (offset 1 → SEMESTER)
        D4 = MOHD KHAIRI    (offset 2 → PENSYARAH)
        D5 = JABATAN ...    (offset 3 → JABATAN)
        D6 = KEMAHIRAN ...  (offset 4 → NAMA KURSUS)
        D7 = GKEB1072       (offset 5 → KOD KURSUS)

    Note: multiple labels can appear on the SAME line, each still getting
    its own sequential offset.
    """
    result: dict[str, str | list[str]] = {}

    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_row), values_only=False):
        for cell in row:
            raw_text = cell.value
            if raw_text is None:
                continue
            text = str(raw_text)

            # Only consider cells with 2+ lines
            if "\n" not in text and text.count(":") < 2:
                continue

            # Scan the FULL text for all anchor keywords in order of appearance
            # Each keyword found gets the next sequential offset
            text_lower = text.lower()
            found_labels: list[tuple[int, str]] = []  # (position_in_text, field_name)

            for anchors, field_name in _METADATA_ANCHORS:
                best_pos = -1
                for anchor in anchors:
                    pos = text_lower.find(anchor)
                    if pos >= 0 and (best_pos < 0 or pos < best_pos):
                        best_pos = pos
                if best_pos >= 0:
                    found_labels.append((best_pos, field_name))

            if len(found_labels) < 3:
                continue  # Not a merged-label block

            # Sort by position in text to get correct sequential order
            found_labels.sort(key=lambda x: x[0])

            # This IS a merged-label cell — find value column
            label_row = cell.row
            label_col = cell.column

            # Find value column: scan right for first non-empty, non-label cell
            value_col = None
            for search_col in range(label_col + 1, min(ws.max_column + 1, label_col + 6)):
                for test_offset in range(min(3, len(found_labels))):
                    try:
                        test_cell = ws.cell(row=label_row + test_offset, column=search_col)
                        if test_cell.value is not None:
                            val = _clean(test_cell.value)
                            if val and not _is_label_cell(val):
                                value_col = search_col
                                break
                    except Exception:
                        pass
                if value_col:
                    break

            if not value_col:
                continue

            # Map each field to its value using SEQUENTIAL offset
            for seq_idx, (_, field_name) in enumerate(found_labels):
                if field_name in result:
                    continue
                value_row = label_row + seq_idx
                try:
                    val_cell = ws.cell(row=value_row, column=value_col)
                    if val_cell.value is not None:
                        value = _clean(val_cell.value)
                        if value:
                            if field_name == "kumpulan_diajar":
                                result[field_name] = _split_kumpulan(value)
                            else:
                                result[field_name] = value
                except Exception:
                    pass

    # --- Fallback: regex-based detection for any remaining critical fields ---
    if "kod_kursus" not in result:
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_row), values_only=False):
            for cell in row:
                if cell.value is None:
                    continue
                m = _KOD_KURSUS_PATTERN.search(str(cell.value))
                if m:
                    code = m.group(1)
                    full_text = str(cell.value).strip()
                    if full_text == code or len(full_text) < 15:
                        result["kod_kursus"] = code
                        break
            if "kod_kursus" in result:
                break

    if "pensyarah" not in result:
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_row), values_only=False):
            for cell in row:
                if cell.value is None:
                    continue
                text = str(cell.value).strip()
                m = _NAMA_PATTERN.search(text)
                if m and text == m.group(0):
                    result["pensyarah"] = _clean(m.group(1))
                    break
            if "pensyarah" in result:
                break

    return result


# ═══════════════════════════════════════════════════════════════════════════
#  XLSX WEEKLY DATA EXTRACTION (anchor-based, no hardcoded columns)
# ═══════════════════════════════════════════════════════════════════════════

def _find_header_row(ws, max_search: int = 25) -> Optional[int]:
    """Find the row containing the 'Minggu' column header.
    Also matches truncated text like 'inggu' from merged cells."""
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_search), values_only=False):
        for cell in row:
            if cell.value is not None:
                val = str(cell.value).strip().lower()
                # Match "Minggu", "Week", or truncated "inggu" from merged cells
                if val in ("minggu", "week") or "minggu" in val or "inggu" in val:
                    return cell.row
    return None


def _map_columns(ws, header_row: int) -> dict[str, int]:
    """
    Dynamically map column positions from the header row.
    Scans header_row ± 2 rows to catch multi-row headers and merged labels.
    Returns dict mapping field names to column indices.
    """
    col_map: dict[str, int] = {}

    header_anchors = {
        "minggu": ["minggu", "week"],
        "tarikh": ["tarikh", "date"],
        "topik": ["topik", "tajuk", "topic"],
        "jam": ["jam", "hour", "interaksi", "bersemuka", "f2f"],
        "catatan": ["catatan", "refleksi", "nota"],
    }

    # Scan wider range: header_row - 2 to header_row + 2
    for scan_row in range(max(1, header_row - 2), min(ws.max_row + 1, header_row + 3)):
        for cell in ws[scan_row]:
            if cell.value is None:
                continue
            header_text = str(cell.value).strip().lower()
            for field_key, anchor_list in header_anchors.items():
                if field_key in col_map:
                    continue
                if any(a in header_text for a in anchor_list):
                    col_map[field_key] = cell.column
                    break

    return col_map


def _discover_jam_columns(ws, header_row: int) -> dict[str, int]:
    """
    Discover K/T/A/L sub-header columns for jam interaksi.
    Scans rows below the header for the characteristic K, T, A, L pattern.
    """
    jam_col_map: dict[str, int] = {}

    for sub_row in range(header_row + 1, min(header_row + 6, ws.max_row + 1)):
        row_cells = list(ws.iter_rows(min_row=sub_row, max_row=sub_row, values_only=False))[0]
        ktal_cells = {}
        for c in row_cells:
            if c.value is not None:
                v = str(c.value).strip()
                if v.upper() in ("K", "T", "A", "L") and len(v) == 1:
                    ktal_cells[c.column] = v.upper()

        if len(ktal_cells) >= 4:
            sorted_cols = sorted(ktal_cells.keys())
            # First group = F2F, second group = ODL
            for col in sorted_cols[:4]:
                jam_col_map[f"{ktal_cells[col]}(F2F)"] = col
            for col in sorted_cols[4:8]:
                jam_col_map[f"{ktal_cells[col]}(ODL)"] = col

            # Look for NF2F column (typically after the K/T/A/L groups)
            max_jam = max(sorted_cols) if sorted_cols else 0
            for scan_r in range(header_row, sub_row + 1):
                for cell in ws[scan_r]:
                    if cell.value is None or cell.column <= max_jam:
                        continue
                    v = str(cell.value).strip().lower()
                    if any(kw in v for kw in ["tidak bersemuka", "nf2f", "non face", "(c)"]):
                        jam_col_map["NF2F"] = cell.column
                        break
            break

    return jam_col_map


def _extract_weeks_from_xlsx(ws, header_row: int, col_map: dict[str, int],
                              jam_col_map: dict[str, int]) -> list[WeekData]:
    """
    Extract weekly data rows from the worksheet.
    Handles merged cells by looking up to 3 rows above for topik values.
    Handles multiline tarikh (e.g. '18 – 22\nOgos 2025').
    """
    weeks: list[WeekData] = []
    minggu_col = col_map.get("minggu", 1)

    # Find first data row (first row with a numeric minggu value)
    data_start = None
    for r in range(header_row + 1, min(header_row + 10, ws.max_row + 1)):
        val = ws.cell(row=r, column=minggu_col).value
        if val is not None:
            try:
                int(val)
                data_start = r
                break
            except (ValueError, TypeError):
                continue

    if not data_start:
        return weeks

    for row_idx in range(data_start, ws.max_row + 1):
        minggu_val = ws.cell(row=row_idx, column=minggu_col).value
        if minggu_val is None:
            continue
        try:
            week_num = int(minggu_val)
        except (ValueError, TypeError):
            continue
        if week_num < 1 or week_num > 19:
            continue

        # --- Tarikh: read at current row, handle multiline ---
        tarikh = ""
        if "tarikh" in col_map:
            tarikh_raw = ws.cell(row=row_idx, column=col_map["tarikh"]).value
            if tarikh_raw is not None:
                tarikh = _clean(tarikh_raw)

        # --- Topik: try current row, then scan UP for merged cells ---
        topik = ""
        if "topik" in col_map:
            topik_col = col_map["topik"]
            topik_raw = ws.cell(row=row_idx, column=topik_col).value
            if topik_raw is not None:
                topik = _clean(topik_raw)
            # If empty, scan up to 3 rows above (merged cell value lives on first row)
            if not topik:
                for scan_up in range(1, 4):
                    check_row = row_idx - scan_up
                    if check_row < data_start:
                        break
                    val = ws.cell(row=check_row, column=topik_col).value
                    if val is not None:
                        candidate = _clean(val)
                        # Make sure it's an actual topik (not a week number or date)
                        if candidate and len(candidate) > 5 and not candidate.isdigit():
                            topik = candidate
                            break

        # --- Catatan ---
        catatan = ""
        if "catatan" in col_map:
            catatan_raw = ws.cell(row=row_idx, column=col_map["catatan"]).value
            if catatan_raw is not None:
                catatan = _clean(catatan_raw)

        # Exception weeks (cuti/peperiksaan/ulangkaji) are PRESERVED with their label.
        # The enrichment step will see the label via is_exception_week() and skip AI generation.
        topik_lower = topik.lower()
        if any(kw in topik_lower for kw in ["cuti", "peperiksaan", "ulangkaji",
                                             "pertengahan semester"]):
            # Normalise to exact exception label if recognisable, else keep raw topik
            if "pertengahan semester" in topik_lower or "cuti" in topik_lower:
                # Distinguish mid-sem vs end-sem by presence of "akhir"
                if "akhir semester" in topik_lower:
                    topik = "CUTI AKHIR SEMESTER IPG"
                else:
                    topik = "CUTI PERTENGAHAN SEMESTER IPG"
            elif "ulangkaji" in topik_lower:
                topik = "MINGGU ULANGKAJI"
            elif "peperiksaan" in topik_lower:
                topik = "PEPERIKSAAN AKHIR"
            # topik is now a clean exception label — NOT blanked

        # Extract jam interaksi
        jam_parts = []
        for label, col_idx in jam_col_map.items():
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                try:
                    num = int(val) if str(val).strip().replace("`", "").isdigit() else 0
                    if num > 0:
                        jam_parts.append(f"{label}:{num}")
                except (ValueError, TypeError):
                    pass

        weeks.append(WeekData(
            minggu=week_num,
            tarikh=_format_tarikh(tarikh),
            topik=topik,
            jam=", ".join(jam_parts) if jam_parts else "",
            hpk="HPK",
            catatan=catatan,
        ))

    return weeks


# ═══════════════════════════════════════════════════════════════════════════
#  MAIN XLSX PARSER (combines Layer 1 + Layer 2)
# ═══════════════════════════════════════════════════════════════════════════

def parse_input_xlsx(file_bytes: bytes) -> tuple[UploadMetadata, list[WeekData]]:
    """
    Parse XLSX with two-layer metadata extraction + dynamic weekly data.
    Layer 1: anchor-based (standard layouts)
    Layer 2: pattern-based (merged-label layouts)
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # --- Metadata extraction ---
    # Layer 1: anchor-based
    meta_dict = _extract_metadata_anchored(ws, max_row=20)

    # Layer 2: label-count offset (handles merged-label layouts).
    # Layer 2 is MORE PRECISE than Layer 1 for non-standard layouts,
    # so its results ALWAYS OVERRIDE Layer 1.
    pattern_meta = _extract_metadata_patterns(ws, max_row=20)
    if pattern_meta:
        for key, val in pattern_meta.items():
            if val:  # Always override if Layer 2 found something
                meta_dict[key] = val

    metadata = UploadMetadata(
        program=str(meta_dict.get("program", "") or ""),
        semester=str(meta_dict.get("semester", "") or ""),
        tahun=str(meta_dict.get("tahun", "") or ""),
        pensyarah=str(meta_dict.get("pensyarah", "") or ""),
        ambilan=str(meta_dict.get("ambilan", "") or ""),
        jabatan=str(meta_dict.get("jabatan", "") or ""),
        kumpulan_diajar=meta_dict.get("kumpulan_diajar", []) if isinstance(meta_dict.get("kumpulan_diajar"), list) else _split_kumpulan(str(meta_dict.get("kumpulan_diajar", ""))),
        nama_kursus=str(meta_dict.get("nama_kursus", "") or ""),
        kod_kursus=str(meta_dict.get("kod_kursus", "") or ""),
        jumlah_kredit=str(meta_dict.get("jumlah_kredit", "") or ""),
    )

    # --- Weekly data extraction ---
    weeks: list[WeekData] = []
    header_row = _find_header_row(ws)
    if header_row:
        col_map = _map_columns(ws, header_row)
        jam_col_map = _discover_jam_columns(ws, header_row)
        weeks = _extract_weeks_from_xlsx(ws, header_row, col_map, jam_col_map)

    if not weeks:
        weeks = [WeekData(minggu=i, hpk="HPK") for i in range(1, 15)]

    weeks.sort(key=lambda w: w.minggu)
    wb.close()

    print(f"[XLSX Parse] nama_kursus='{metadata.nama_kursus}', kod='{metadata.kod_kursus}', "
          f"pensyarah='{metadata.pensyarah}', kumpulan={metadata.kumpulan_diajar}, weeks={len(weeks)}")

    return metadata, weeks


# ═══════════════════════════════════════════════════════════════════════════
#  PDF PARSER (regex + keyword proximity)
# ═══════════════════════════════════════════════════════════════════════════

def parse_input_pdf(file_bytes: bytes) -> tuple[UploadMetadata, list[WeekData]]:
    """
    Parse PDF using keyword-proximity extraction.
    For scanned/image PDFs with no extractable text, returns empty metadata
    (the AI hybrid parser will handle those via Gemini vision).
    """
    metadata = UploadMetadata()
    weeks: list[WeekData] = []

    try:
        pdf = pdfplumber.open(BytesIO(file_bytes))
    except Exception:
        return metadata, [WeekData(minggu=i, hpk="HPK") for i in range(1, 15)]

    # Extract all text
    full_text = ""
    for page in pdf.pages:
        page_text = page.extract_text() or ""
        full_text += page_text + "\n"
    pdf.close()

    if len(full_text.strip()) < 50:
        # Likely a scanned PDF — no text extracted
        return metadata, [WeekData(minggu=i, hpk="HPK") for i in range(1, 15)]

    lines = [ln.strip() for ln in full_text.split("\n") if ln.strip()]

    # --- Metadata from text ---
    def _find_value_after_anchor(anchor: str) -> str:
        """Find the value after an anchor word in the text."""
        for line in lines:
            idx = line.lower().find(anchor.lower())
            if idx == -1:
                continue
            # Get text after the anchor
            after = line[idx + len(anchor):]
            # Remove leading colon/whitespace
            after = after.lstrip(" \t:")
            if after:
                # Take until end of line or until next anchor keyword
                return _clean(after)
        return ""

    # Extract each metadata field
    for anchors, field_name in _METADATA_ANCHORS:
        for anchor in anchors:
            value = _find_value_after_anchor(anchor)
            if value:
                # Trim at known anchor boundaries
                # e.g., "PSIKOLOGI PEMBELAJARAN JUMLAH KREDIT: 3" → stop before "JUMLAH"
                for stop_anchors, _ in _METADATA_ANCHORS:
                    for stop in stop_anchors:
                        stop_idx = value.lower().find(stop)
                        if stop_idx > 0:
                            value = value[:stop_idx].strip().rstrip(":")
                            break

                if not value:
                    continue

                if field_name == "kumpulan_diajar":
                    if not metadata.kumpulan_diajar:
                        metadata.kumpulan_diajar = _split_kumpulan(value)
                elif not getattr(metadata, field_name, None):
                    setattr(metadata, field_name, value)
                break

    # Fallback: pattern-based detection on full text
    if not metadata.kod_kursus:
        m = _KOD_KURSUS_PATTERN.search(full_text[:2000])
        if m:
            metadata.kod_kursus = m.group(1)

    if not metadata.pensyarah:
        m = _NAMA_PATTERN.search(full_text[:2000])
        if m:
            metadata.pensyarah = _clean(m.group(1))

    # --- Weekly data from text ---
    seen_weeks: set[int] = set()
    # Try table-like extraction first
    all_tables = []
    try:
        pdf2 = pdfplumber.open(BytesIO(file_bytes))
        for page in pdf2.pages[:5]:
            tables = page.extract_tables()
            if tables:
                all_tables.extend(tables)
        pdf2.close()
    except Exception:
        pass

    for table in all_tables:
        if not table or len(table) < 2:
            continue
        for row in table:
            if not row or len(row) < 3:
                continue
            # First cell might be week number
            try:
                week_num = int(str(row[0]).strip())
            except (ValueError, TypeError):
                continue
            if week_num < 1 or week_num > 20 or week_num in seen_weeks:
                continue
            # Find topik (longest cell in the row)
            topik = ""
            tarikh = ""
            for cell in row[1:]:
                cell_str = _clean(cell)
                if not cell_str:
                    continue
                # Date-like cell
                if re.search(r'\d{1,2}\s*[-–]\s*\d{1,2}', cell_str) and not tarikh:
                    tarikh = cell_str
                elif len(cell_str) > len(topik):
                    topik = cell_str

            topik_lower = topik.lower()
            if any(kw in topik_lower for kw in ["cuti", "peperiksaan", "ulangkaji",
                                                  "pertengahan semester"]):
                # Normalise to exact exception label
                if "akhir semester" in topik_lower:
                    topik = "CUTI AKHIR SEMESTER IPG"
                elif "pertengahan" in topik_lower or "cuti" in topik_lower:
                    topik = "CUTI PERTENGAHAN SEMESTER IPG"
                elif "ulangkaji" in topik_lower:
                    topik = "MINGGU ULANGKAJI"
                elif "peperiksaan" in topik_lower:
                    topik = "PEPERIKSAAN AKHIR"
            # Skip rows where topik is actually just a date range (spurious weeks)
            if re.match(r'^\d{1,2}\s+\w+\s*[-–]\s*\d{1,2}\s+\w+\s+\d{4}$', topik.strip()):
                continue

            seen_weeks.add(week_num)
            weeks.append(WeekData(
                minggu=week_num,
                tarikh=_format_tarikh(tarikh),
                topik=topik,
                hpk="HPK",
            ))

    if not weeks:
        weeks = [WeekData(minggu=i, hpk="HPK") for i in range(1, 15)]

    weeks.sort(key=lambda w: w.minggu)
    return metadata, weeks


# ---------------------------------------------------------------------------
#  Tarikh gap-fill: interpolate missing dates from known neighbors
# ---------------------------------------------------------------------------

_MALAY_MONTHS = {
    "januari": 1, "februari": 2, "mac": 3, "april": 4, "mei": 5, "jun": 6,
    "julai": 7, "ogos": 8, "september": 9, "oktober": 10, "november": 11, "disember": 12,
}

def _parse_tarikh_to_date(tarikh: str):
    """Try to parse '18 Ogos 2025' into a datetime.date, else None."""
    import datetime
    parts = tarikh.strip().split()
    if len(parts) < 3:
        return None
    try:
        day = int(parts[0])
        month = _MALAY_MONTHS.get(parts[1].lower())
        year = int(parts[2])
        if month:
            return datetime.date(year, month, day)
    except (ValueError, IndexError):
        pass
    return None


def _date_to_tarikh(d) -> str:
    """Convert datetime.date back to '18 Ogos 2025' format."""
    month_names = {
        1: "Januari", 2: "Februari", 3: "Mac", 4: "April", 5: "Mei", 6: "Jun",
        7: "Julai", 8: "Ogos", 9: "September", 10: "Oktober", 11: "November", 12: "Disember",
    }
    return f"{d.day} {month_names[d.month]} {d.year}"


def _fill_tarikh_gaps(weeks: list[WeekData]) -> None:
    """Fill empty tarikh values by interpolating from the nearest known date."""
    import datetime
    if not weeks:
        return

    # Build map of known dates
    known: dict[int, object] = {}
    for w in weeks:
        if w.tarikh:
            d = _parse_tarikh_to_date(w.tarikh)
            if d:
                known[w.minggu] = d

    if not known:
        return

    # For each week with missing tarikh, extrapolate from nearest known
    for w in weeks:
        if w.tarikh:
            continue
        # Find closest known week
        closest_week = min(known.keys(), key=lambda k: abs(k - w.minggu))
        closest_date = known[closest_week]
        delta_weeks = w.minggu - closest_week
        estimated = closest_date + datetime.timedelta(weeks=delta_weeks)
        w.tarikh = _date_to_tarikh(estimated)
        print(f"[Tarikh Fill] W{w.minggu}: interpolated from W{closest_week} → {w.tarikh}")


# ═══════════════════════════════════════════════════════════════════════════
#  LAYER 3: LLM-ASSISTED HYBRID PARSER
# ═══════════════════════════════════════════════════════════════════════════

async def parse_with_ai(file_bytes: bytes, filename: str) -> tuple[UploadMetadata, list[WeekData]]:
    """
    Three-layer hybrid parser:
      1. AI (Gemini) as PRIMARY extractor — handles any layout
      2. Deterministic anchor/pattern parsing as VERIFICATION overlay
      3. Pattern regex as final fallback for critical fields

    The AI result is trusted for topik/tarikh/weeks.
    Deterministic results override AI for nama_kursus, kod_kursus, pensyarah, kumpulan
    (because these can be extracted precisely from structured cells).
    """
    filename_lower = filename.lower()

    # Step 1: Extract text for AI
    if filename_lower.endswith(".pdf"):
        raw_text = _extract_text_from_pdf(file_bytes)
    else:
        raw_text = _extract_text_from_xlsx(file_bytes)

    if not raw_text or len(raw_text.strip()) < 50:
        # Scanned PDF or nearly empty file — fall back to deterministic parser
        print(f"[AI Parse] Text too short ({len((raw_text or '').strip())} chars), falling back to deterministic parser")
        if filename_lower.endswith(".pdf"):
            return parse_input_pdf(file_bytes)
        else:
            return parse_input_xlsx(file_bytes)

    # Step 2: AI extraction
    data = await extract_file_content(raw_text)
    meta_raw = data.get("metadata", {}) or {}

    # Normalise keys
    meta_dict: dict[str, str] = {}
    for k, v in meta_raw.items():
        if isinstance(k, str):
            meta_dict[k.strip().lower().replace(" ", "_")] = v

    def _ai_get(keys: list[str], substrings: list[str] | None = None, default=""):
        for key in keys:
            kn = key.strip().lower().replace(" ", "_")
            if kn in meta_dict and meta_dict[kn] not in (None, ""):
                return meta_dict[kn]
        if substrings:
            for k, v in meta_dict.items():
                if all(s in k.lower() for s in substrings) and v not in (None, ""):
                    return v
        return default

    raw_kumpulan = _ai_get(
        ["kumpulan_diajar", "kumpulan", "kumpulan_ajar"],
        ["kumpulan"], default=[],
    )
    if isinstance(raw_kumpulan, str):
        kumpulan_list = _split_kumpulan(raw_kumpulan)
    elif isinstance(raw_kumpulan, list):
        kumpulan_list = [str(g).strip() for g in raw_kumpulan if str(g).strip()]
    else:
        kumpulan_list = []

    metadata = UploadMetadata(
        program=str(_ai_get(["program"], ["program"]) or ""),
        semester=str(_ai_get(["semester"], ["semester"]) or ""),
        tahun=str(_ai_get(["tahun", "tahun_akademik"], ["tahun"]) or ""),
        pensyarah=str(_ai_get(["pensyarah", "nama_pensyarah"], ["pensyarah"]) or ""),
        ambilan=str(_ai_get(["ambilan", "intake"], ["ambilan"]) or ""),
        jabatan=str(_ai_get(["jabatan", "unit"], ["jabatan"]) or ""),
        kumpulan_diajar=kumpulan_list,
        nama_kursus=str(_ai_get(["nama_kursus", "nama_mata_pelajaran"], ["kursus"]) or ""),
        kod_kursus=str(_ai_get(["kod_kursus", "kod"], ["kod"]) or ""),
        jumlah_kredit=str(_ai_get(["jumlah_kredit", "kredit"], ["kredit"]) or ""),
    )

    print(f"[AI Parse] AI extracted: nama_kursus='{metadata.nama_kursus}', "
          f"kod='{metadata.kod_kursus}', pensyarah='{metadata.pensyarah}'")

    # Step 3: Extract weeks from AI
    weeks: list[WeekData] = []
    for w in data.get("weeks", []):
        try:
            minggu = int(w.get("minggu", 0))
            if minggu < 1 or minggu > 19:
                continue
            raw_topik = str(w.get("topik", ""))
            # Normalise exception-week labels from AI output to canonical form
            raw_lower = raw_topik.strip().lower()
            if any(kw in raw_lower for kw in ["cuti", "peperiksaan", "ulangkaji", "pertengahan semester"]):
                if "akhir semester" in raw_lower:
                    raw_topik = "CUTI AKHIR SEMESTER IPG"
                elif "pertengahan" in raw_lower or ("cuti" in raw_lower and "akhir" not in raw_lower):
                    raw_topik = "CUTI PERTENGAHAN SEMESTER IPG"
                elif "ulangkaji" in raw_lower:
                    raw_topik = "MINGGU ULANGKAJI"
                elif "peperiksaan" in raw_lower:
                    raw_topik = "PEPERIKSAAN AKHIR"
            weeks.append(WeekData(
                minggu=minggu,
                tarikh=_format_tarikh(str(w.get("tarikh", ""))),
                topik=raw_topik,
                jam="",
                hpk="HPK",
                catatan=str(w.get("catatan", "")),
            ))
        except (ValueError, TypeError):
            continue

    if not weeks:
        raise ValueError("AI tidak berjaya mengekstrak data minggu")

    weeks.sort(key=lambda w: w.minggu)

    # Step 4: Deterministic overlay — XLSX gets anchor+pattern extraction
    if not filename_lower.endswith(".pdf"):
        try:
            det_meta, det_weeks = parse_input_xlsx(file_bytes)

            # Jam interaksi: always from deterministic parser
            jam_map = {rw.minggu: rw.jam for rw in det_weeks if rw.jam}
            for week in weeks:
                if week.minggu in jam_map:
                    week.jam = jam_map[week.minggu]

            # Critical fields: deterministic overrides AI
            if det_meta.nama_kursus:
                metadata.nama_kursus = det_meta.nama_kursus
            if det_meta.kod_kursus:
                metadata.kod_kursus = det_meta.kod_kursus
            if det_meta.pensyarah:
                metadata.pensyarah = det_meta.pensyarah
            if det_meta.kumpulan_diajar:
                metadata.kumpulan_diajar = det_meta.kumpulan_diajar

            # Fill remaining gaps
            for field in ["program", "semester", "tahun", "jabatan", "ambilan", "jumlah_kredit"]:
                if not getattr(metadata, field) and getattr(det_meta, field, None):
                    setattr(metadata, field, getattr(det_meta, field))

            # Fill empty tarikh from deterministic
            tarikh_map = {rw.minggu: rw.tarikh for rw in det_weeks if rw.tarikh}
            for week in weeks:
                if not week.tarikh and week.minggu in tarikh_map:
                    week.tarikh = tarikh_map[week.minggu]

        except Exception as e:
            print(f"[Hybrid] Deterministic XLSX overlay failed: {e}")
    else:
        try:
            det_meta, _ = parse_input_pdf(file_bytes)
            if det_meta.nama_kursus:
                metadata.nama_kursus = det_meta.nama_kursus
            if det_meta.kod_kursus:
                metadata.kod_kursus = det_meta.kod_kursus
            if det_meta.pensyarah:
                metadata.pensyarah = det_meta.pensyarah
            if det_meta.kumpulan_diajar:
                metadata.kumpulan_diajar = det_meta.kumpulan_diajar
        except Exception as e:
            print(f"[Hybrid] Deterministic PDF overlay failed: {e}")

    # Fill tarikh gaps for PDF — interpolate from neighbors
    _fill_tarikh_gaps(weeks)

    print(f"[AI Parse] Final: nama_kursus='{metadata.nama_kursus}', "
          f"kod='{metadata.kod_kursus}', pensyarah='{metadata.pensyarah}', "
          f"kumpulan={metadata.kumpulan_diajar}, weeks={len(weeks)}")

    return metadata, weeks


# ═══════════════════════════════════════════════════════════════════════════
#  TEXT EXTRACTION HELPERS (for AI / Gemini input)
# ═══════════════════════════════════════════════════════════════════════════

def _extract_text_from_xlsx(file_bytes: bytes) -> str:
    """Convert XLSX to readable text for AI. No hardcoded coordinates."""
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    lines: list[str] = []

    # Find where weekly data starts
    weekly_start = ws.max_row + 1
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 25), values_only=False):
        for cell in row:
            if cell.value and "minggu" in str(cell.value).strip().lower():
                weekly_start = cell.row
                break
        if weekly_start <= ws.max_row:
            break

    # Metadata area
    lines.append("=== MAKLUMAT KURSUS (METADATA) ===")
    for row in ws.iter_rows(min_row=1, max_row=min(weekly_start - 1, ws.max_row), values_only=False):
        cells = [str(c.value).strip() for c in row if c.value is not None and str(c.value).strip()]
        if cells:
            lines.append(" : ".join(cells))

    # Weekly data area
    lines.append("\n=== DATA MINGGUAN ===")
    for row in ws.iter_rows(min_row=weekly_start, values_only=False):
        cells = [str(c.value).strip() for c in row if c.value is not None and str(c.value).strip()]
        if cells:
            lines.append(" | ".join(cells))

    wb.close()
    return "\n".join(lines)


def _extract_text_from_pdf(file_bytes: bytes) -> str:
    """Extract all text from PDF for AI processing."""
    try:
        pdf = pdfplumber.open(BytesIO(file_bytes))
        parts = [page.extract_text() or "" for page in pdf.pages]
        pdf.close()
        return "\n".join(parts)
    except Exception:
        return ""
