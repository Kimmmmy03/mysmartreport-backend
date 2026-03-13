"""
MySmartReport — Excel Generator Service

Each week loads a FRESH copy of template.xlsx and pastes data into cells.
No manual formatting, no structure copying — template has everything.

Cell mapping:
  C3=program, L3=kod kursus, C4=semester, G4=tahun, L4=jumlah kredit
  C5=nama kursus, L5=kumpulan diajar, C6=minggu, G6=tarikh
  B11=Topik, C11=Hasil Pembelajaran, F11=HPK
  G11=Strategi Kuliah (rows 11-16 merged), G17=Strategi Tutorial (rows 17-21 merged)
  H-P row 11/17=Jam Interaksi, Q11=Refleksi Kuliah, Q17=Refleksi Tutorial
"""

import re
from copy import copy
from io import BytesIO
from zipfile import ZipFile
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from config import settings
from models import SessionDraft, WeekData, GroupAttendance


def _sanitize_for_excel(value):
    """
    Sanitize a value before writing to an Excel cell.
    Strips illegal XML 1.0 control characters that cause
    'Repaired Part: /xl/worksheets/sheet1.xml' errors.

    XML 1.0 legal chars: #x9 | #xA | #xD | [#x20-#xD7FF] | [#xE000-#xFFFD]
    Everything else (0x00-0x08, 0x0B, 0x0C, 0x0E-0x1F, 0xFFFE, 0xFFFF) is illegal.
    """
    if value is None or not isinstance(value, str):
        return value
    # Remove illegal XML control characters
    return re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F\uFFFE\uFFFF]', '', value)


def _estimate_row_height(text, col_width_chars=40, line_height_pt=14.5):
    """Estimate row height (pt) needed for text with word wrap enabled."""
    if not text or not isinstance(text, str):
        return 0
    total_visual_lines = 0
    for line in text.split('\n'):
        if not line.strip():
            total_visual_lines += 1
        else:
            total_visual_lines += max(1, -(-len(line) // col_width_chars))
    return total_visual_lines * line_height_pt


def _bold_labels_in_shared_strings(xml_text, target_indices):
    """
    Convert specific shared string entries to rich text with bold labels.
    target_indices: {shared_string_index: cell_ref}
    """
    cell_labels = {
        'G11': ['Kuliah'],
        'G17': ['Tutorial'],
        'G22': ['E-pembelajaran'],
        'Q11': ['Catatan:', 'Refleksi:'],
        'Q17': ['Refleksi:'],
        'Q22': ['Refleksi:'],
    }

    # Process in reverse index order to preserve string positions
    for idx in sorted(target_indices.keys(), reverse=True):
        cell_ref = target_indices[idx]
        labels = cell_labels.get(cell_ref, [])
        if not labels:
            continue

        # Find the Nth <si> entry (0-indexed)
        si_pattern = re.compile(r'<si>(.*?)</si>', re.DOTALL)
        matches = list(si_pattern.finditer(xml_text))
        if idx >= len(matches):
            continue

        si_match = matches[idx]
        inner = si_match.group(1)

        # Only process plain text entries (not already rich text)
        if '<r>' in inner:
            continue

        # Extract text from <t> element
        t_match = re.search(r'<t(?:\s[^>]*)?>(.*?)</t>', inner, re.DOTALL)
        if not t_match:
            continue

        text = t_match.group(1)

        # Split at label positions
        pattern = '(' + '|'.join(re.escape(l) for l in labels) + ')'
        parts = re.split(pattern, text)

        if len(parts) <= 1:
            continue

        # Build rich text runs
        runs = []
        for part in parts:
            if not part:
                continue
            if part in labels:
                runs.append(f'<r><rPr><b/></rPr><t xml:space="preserve">{part}</t></r>')
            else:
                runs.append(f'<r><t xml:space="preserve">{part}</t></r>')

        new_si = '<si>' + ''.join(runs) + '</si>'
        xml_text = xml_text[:si_match.start()] + new_si + xml_text[si_match.end():]

    return xml_text


def _fix_xlsx(buffer: BytesIO) -> BytesIO:
    """
    Post-process an xlsx buffer to fix openpyxl's invalid XML that
    causes Excel to show 'Repaired Part: sheet1.xml'.

    Fixes:
    - Floating point row heights (96.90000000000001 → 96.9)
    - Empty cells with t="n" (invalid number type with no value)
    - _x000a_ encoded newlines in headers (should be real newlines)
    - Empty evenHeader/evenFooter/firstHeader/firstFooter elements
    - outlinePr element not in original template
    - baseColWidth attribute not in original template
    - style attribute on col elements not in original template
    - Illegal XML control characters in cell data
    - Bold labels (Catatan:, Refleksi:, Kuliah, Tutorial) via rich text
    - firstSheet="1" in workbook.xml (invalid for single-sheet files)
    """
    def _round_row_ht(m):
        prefix, val, rest = m.group(1), m.group(2), m.group(3)
        try:
            return f'{prefix}ht="{round(float(val), 2):g}"{rest}'
        except ValueError:
            return m.group(0)

    buffer.seek(0)

    # Read all zip entries
    entries = []  # list of (ZipInfo, bytes)
    with ZipFile(buffer, 'r') as zf:
        for item in zf.infolist():
            entries.append((item, zf.read(item.filename)))

    # Track shared string indices for bold-label cells
    bold_targets = {}  # {shared_string_index: cell_ref}

    # Process each entry
    processed = []
    for item, data in entries:
        if item.filename.startswith('xl/worksheets/sheet') and item.filename.endswith('.xml'):
            text = data.decode('utf-8')

            # Find shared string indices for cells that need bold labels
            for cell_ref in ['G11', 'G17', 'G22', 'Q11', 'Q17', 'Q22']:
                m = re.search(
                    rf'<c\s(?=[^>]*r="{cell_ref}")(?=[^>]*t="s")[^>]*><v>(\d+)</v>',
                    text
                )
                if m:
                    bold_targets[int(m.group(1))] = cell_ref

            # Fix floating point row heights
            text = re.sub(r'(<row [^>]*?)ht="([^"]+)"([^>]*>)', _round_row_ht, text)
            # Fix empty cells with t="n"
            text = re.sub(r'(<c r="[A-Z]+\d+" s="\d+") t="n" />', r'\1 />', text)
            # Fix _x000a_ → real newline in headers
            text = text.replace('_x000a_', '\n')
            # Remove empty header/footer elements
            text = text.replace('<evenHeader />', '')
            text = text.replace('<evenFooter />', '')
            text = text.replace('<firstHeader />', '')
            text = text.replace('<firstFooter />', '')
            # Remove outlinePr element
            text = re.sub(r'<outlinePr[^/]*/>\s*', '', text)
            # Remove baseColWidth from sheetFormatPr
            text = re.sub(r' baseColWidth="\d+"', '', text)
            # Remove style attribute from col elements
            text = re.sub(r' style="\d+"', '', text)
            # Strip illegal XML control characters from cell data
            text = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', text)

            data = text.encode('utf-8')

        elif item.filename.endswith('workbook.xml') and 'rels' not in item.filename:
            text = data.decode('utf-8')
            # Fix firstSheet for single-sheet workbooks
            text = re.sub(r'firstSheet="\d+"', 'firstSheet="0"', text)
            data = text.encode('utf-8')

        processed.append((item, data))

    # Apply bold labels to sharedStrings.xml
    if bold_targets:
        for i, (item, data) in enumerate(processed):
            if item.filename == 'xl/sharedStrings.xml':
                text = data.decode('utf-8')
                text = _bold_labels_in_shared_strings(text, bold_targets)
                processed[i] = (item, text.encode('utf-8'))
                break

    # Write output zip
    out = BytesIO()
    with ZipFile(out, 'w') as zf_out:
        for item, data in processed:
            zf_out.writestr(item, data)
    out.seek(0)
    return out


def _setup_content_merges(ws):
    """
    Restructure content area: rows 11-13 for Kuliah, rows 14-16 for Tutorial,
    rows 17-19 for E-pembelajaran.
    3 rows per section (max 3 combined rows as requested).
    Unmerges old template merges, creates new vertical merges, copies styles.
    """
    from openpyxl.styles import Alignment

    # --- Unmerge existing merges that overlap rows 11-30 ---
    to_unmerge = []
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row >= 11 and merged_range.max_row <= 30:
            to_unmerge.append(str(merged_range))
    for mr in to_unmerge:
        ws.unmerge_cells(mr)

    # --- Copy styles from row 11 to rows 12-19 for key columns ---
    style_cols = ['B', 'C', 'F', 'G', 'Q']
    for col_idx in range(8, 17):  # H=8 to P=16
        style_cols.append(get_column_letter(col_idx))

    for col_letter in style_cols:
        src_cell = ws[f'{col_letter}11']
        for row in range(12, 20):  # rows 12-19
            dst_cell = ws[f'{col_letter}{row}']
            if src_cell.has_style:
                dst_cell.font = copy(src_cell.font)
                dst_cell.border = copy(src_cell.border)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.number_format = src_cell.number_format
                dst_cell.protection = copy(src_cell.protection)
                dst_cell.alignment = copy(src_cell.alignment)

    # --- Kuliah merges: rows 11-13 (3 rows) ---
    ws.merge_cells('B11:B13')
    ws.merge_cells('C11:E13')
    ws.merge_cells('F11:F13')
    ws.merge_cells('G11:G13')
    ws.merge_cells('Q11:Q13')
    for col_idx in range(8, 17):
        col_l = get_column_letter(col_idx)
        ws.merge_cells(f'{col_l}11:{col_l}13')

    # --- Tutorial merges: rows 14-16 (3 rows) ---
    ws.merge_cells('B14:B16')
    ws.merge_cells('C14:E16')
    ws.merge_cells('F14:F16')
    ws.merge_cells('G14:G16')
    ws.merge_cells('Q14:Q16')
    for col_idx in range(8, 17):
        col_l = get_column_letter(col_idx)
        ws.merge_cells(f'{col_l}14:{col_l}16')

    # --- E-pembelajaran merges: rows 17-19 (3 rows) ---
    ws.merge_cells('B17:B19')
    ws.merge_cells('C17:E19')
    ws.merge_cells('F17:F19')
    ws.merge_cells('G17:G19')
    ws.merge_cells('Q17:Q19')
    for col_idx in range(8, 17):
        col_l = get_column_letter(col_idx)
        ws.merge_cells(f'{col_l}17:{col_l}19')

    # --- Ensure wrapText on key cells ---
    for cell_ref in ['B11', 'C11', 'G11', 'Q11', 'G14', 'Q14', 'G17', 'Q17']:
        cell = ws[cell_ref]
        cell.alignment = Alignment(
            horizontal=cell.alignment.horizontal,
            vertical=cell.alignment.vertical or 'top',
            wrap_text=True,
        )


def _load_template_sheet():
    """Load template workbook and return (wb, ws)."""
    wb = load_workbook(settings.TEMPLATE_PATH)
    if "PT03_02_Mingguan_M1" in wb.sheetnames:
        ws = wb["PT03_02_Mingguan_M1"]
    else:
        ws = wb.active
    return wb, ws


def _inject_week_data(ws, week: WeekData, draft: SessionDraft, group: GroupAttendance | None = None):
    """Paste data into template cells. No formatting — template handles it all."""

    s = _sanitize_for_excel  # shorthand

    # Header metadata
    ws["C3"] = s(draft.metadata.program)
    ws["L3"] = s(draft.metadata.kod_kursus)
    ws["C4"] = s(draft.metadata.semester)
    ws["G4"] = s(draft.metadata.tahun)
    ws["L4"] = s(draft.metadata.jumlah_kredit)
    ws["C5"] = s(draft.metadata.nama_kursus)

    # Kumpulan
    if group:
        ws["L5"] = s(group.nama)
    elif draft.kumpulan_list:
        ws["L5"] = s(", ".join(g.nama for g in draft.kumpulan_list))
    elif draft.metadata.kumpulan_diajar:
        ws["L5"] = s(", ".join(draft.metadata.kumpulan_diajar))

    # Week info
    ws["C6"] = week.minggu
    ws["G6"] = s(week.tarikh)

    # Content
    ws["B11"] = s(week.topik)
    ws["C11"] = s(week.hasil_pembelajaran)
    ws["F11"] = s(week.hpk)

    # Strategi — split into Kuliah (G11), Tutorial (G17), E-pembelajaran (G22)
    strategi = week.strategi_aktiviti
    has_tutorial = False
    has_epembelajaran = False
    if strategi:
        lines = strategi.split("\n")
        kuliah_lines = []
        tutorial_lines = []
        epembelajaran_lines = []
        section = "kuliah"
        for line in lines:
            line_lower = line.strip().lower()
            if line_lower.startswith("tutorial"):
                section = "tutorial"
                has_tutorial = True
            elif line_lower.startswith("e-pembelajaran"):
                section = "epembelajaran"
                has_epembelajaran = True
            if section == "kuliah":
                kuliah_lines.append(line)
            elif section == "tutorial":
                tutorial_lines.append(line)
            else:
                epembelajaran_lines.append(line)

        ws["G11"] = s("\n".join(kuliah_lines).strip())
        if tutorial_lines:
            ws["G14"] = s("\n".join(tutorial_lines).strip())
        if epembelajaran_lines:
            ws["G17"] = s("\n".join(epembelajaran_lines).strip())
    else:
        ws["G11"] = s(strategi)

    # Refleksi Kuliah (Q11) — with kehadiran
    catatan_kuliah = []
    if group:
        catatan_kuliah.append(f"Catatan:\nKehadiran pelajar {group.kehadiran}/{group.jumlah_pelajar} orang.")
    elif draft.kumpulan_list:
        kehadiran_text = "Kehadiran pelajar " + ", ".join(
            f"{g.kehadiran}/{g.jumlah_pelajar} orang" for g in draft.kumpulan_list
        )
        catatan_kuliah.append(f"Catatan:\n{kehadiran_text}")
    if week.refleksi:
        catatan_kuliah.append(f"Refleksi:\n{week.refleksi}")
    ws["Q11"] = s("\n\n".join(catatan_kuliah)) if catatan_kuliah else ""

    # Refleksi Tutorial (Q14)
    ws["Q14"] = ""
    if week.refleksi_tutorial:
        ws["Q14"] = s(f"Refleksi:\n{week.refleksi_tutorial}")
    elif has_tutorial and week.refleksi:
        ws["Q14"] = s(f"Refleksi:\n{week.refleksi}")

    # Refleksi E-pembelajaran (Q17)
    ws["Q17"] = ""
    if week.refleksi_epembelajaran:
        ws["Q17"] = s(f"Refleksi:\n{week.refleksi_epembelajaran}")
    elif has_epembelajaran and week.refleksi:
        ws["Q17"] = s(f"Refleksi:\n{week.refleksi}")

    # Jam Interaksi
    kuliah_jam = {
        "K(F2F)": "H11", "A(F2F)": "J11", "L(F2F)": "K11",
        "K(ODL)": "L11", "A(ODL)": "N11", "L(ODL)": "O11",
        "NF2F": "P11",
    }
    tutorial_jam = {
        "T(F2F)": "I14" if has_tutorial else "I11",
        "T(ODL)": "M14" if has_tutorial else "M11",
    }
    if week.jam:
        for part in week.jam.split(","):
            part = part.strip()
            if ":" in part:
                key, val = part.rsplit(":", 1)
                key = key.strip()
                cell_ref = kuliah_jam.get(key) or tutorial_jam.get(key)
                if cell_ref:
                    try:
                        ws[cell_ref] = int(val.strip())
                    except ValueError:
                        pass

    # --- Set font size 10 for all AI content cells ---
    ai_content_font = Font(name="Arial", size=10)
    for cell_ref in ["B11", "C11", "F11", "G11", "G14", "G17", "Q11", "Q14", "Q17"]:
        cell = ws[cell_ref]
        if cell.value:
            cell.font = ai_content_font

    # Setup merged content area (rows 11-13 Kuliah, rows 14-16 Tutorial, rows 17-19 E-pembelajaran)
    _setup_content_merges(ws)

    # Dynamic row heights — distribute across merged rows (3 per section)
    # Column width estimates (chars): B~20, C-E~40, G~55, Q~30
    kuliah_content_height = max(200, max(
        _estimate_row_height(str(ws['B11'].value or ''), 20),
        _estimate_row_height(str(ws['C11'].value or ''), 40),
        _estimate_row_height(str(ws['G11'].value or ''), 55),
        _estimate_row_height(str(ws['Q11'].value or ''), 30),
    ))
    tutorial_content_height = max(120, max(
        _estimate_row_height(str(ws['G14'].value or ''), 55),
        _estimate_row_height(str(ws['Q14'].value or ''), 30),
    ))
    epembelajaran_content_height = max(120, max(
        _estimate_row_height(str(ws['G17'].value or ''), 55),
        _estimate_row_height(str(ws['Q17'].value or ''), 30),
    ))

    # Distribute Kuliah height across rows 11-13 (3 rows, capped at 409.5 each)
    kuliah_per_row = min(409.5, kuliah_content_height / 3)
    for r in range(11, 14):
        ws.row_dimensions[r].height = kuliah_per_row

    # Distribute Tutorial height across rows 14-16 (3 rows)
    tutorial_per_row = min(409.5, tutorial_content_height / 3)
    for r in range(14, 17):
        ws.row_dimensions[r].height = tutorial_per_row

    # Distribute E-pembelajaran height across rows 17-19 (3 rows)
    epembelajaran_per_row = min(409.5, epembelajaran_content_height / 3)
    for r in range(17, 20):
        ws.row_dimensions[r].height = epembelajaran_per_row


def _generate_single_week_buffer(week: WeekData, draft: SessionDraft, group: GroupAttendance | None = None) -> BytesIO:
    """Generate one week Excel by loading a fresh template and pasting data."""
    wb, ws = _load_template_sheet()
    _inject_week_data(ws, week, draft, group)
    ws.title = f"PT03_02_Mingguan_M{week.minggu}"

    buffer = BytesIO()
    wb.save(buffer)
    wb.close()
    return _fix_xlsx(buffer)


def generate_single_excel(draft: SessionDraft, selected_weeks: list[int], input_bytes: bytes | None = None) -> BytesIO:
    """
    Generate a single Excel workbook with one sheet per selected week.
    Each sheet is a fresh template copy with data injected.
    """
    group = draft.kumpulan_list[0] if draft.kumpulan_list else None
    weeks_to_export = sorted(
        [w for w in draft.weeks if w.minggu in selected_weeks],
        key=lambda w: w.minggu,
    )

    if not weeks_to_export:
        from openpyxl import Workbook
        wb = Workbook()
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    # Single week — just load template, inject, save
    if len(weeks_to_export) == 1 and not input_bytes:
        return _generate_single_week_buffer(weeks_to_export[0], draft, group)

    # Multiple weeks — use first week as base, add more sheets from fresh templates
    # Start with input file or first week's template
    if input_bytes:
        base_wb = load_workbook(BytesIO(input_bytes))
    else:
        base_wb, base_ws = _load_template_sheet()
        _inject_week_data(base_ws, weeks_to_export[0], draft, group)
        base_ws.title = f"RPP_Mingguan_{draft.metadata.kod_kursus or 'RPP'} - Minggu {weeks_to_export[0].minggu}"

    # For remaining weeks (or all if input_bytes), generate separate files
    # and copy sheets into the base workbook
    start_idx = 0 if input_bytes else 1
    for week in weeks_to_export[start_idx:]:
        # Generate this week as a standalone buffer
        week_buf = _generate_single_week_buffer(week, draft, group)
        week_wb = load_workbook(week_buf)
        week_ws = week_wb.active

        sheet_title = f"RPP_M{week.minggu}_{draft.metadata.kod_kursus or 'RPP'}"
        new_ws = base_wb.create_sheet(title=sheet_title)

        # Copy all cells with styles
        for row in week_ws.iter_rows():
            for cell in row:
                new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

        # Copy merged cells
        for merged_range in week_ws.merged_cells.ranges:
            new_ws.merge_cells(str(merged_range))

        # Copy column widths and row heights
        for col_letter, dim in week_ws.column_dimensions.items():
            new_ws.column_dimensions[col_letter].width = dim.width
        for row_num, dim in week_ws.row_dimensions.items():
            if dim.height:
                new_ws.row_dimensions[row_num].height = dim.height

        week_wb.close()

    buffer = BytesIO()
    base_wb.save(buffer)
    base_wb.close()
    return _fix_xlsx(buffer)


def generate_per_group_zips(draft: SessionDraft, selected_weeks: list[int]) -> dict[str, BytesIO]:
    """Generate one ZIP per kumpulan group, each containing all selected weeks."""
    result: dict[str, BytesIO] = {}
    groups = draft.kumpulan_list if draft.kumpulan_list else [None]

    for group in groups:
        zip_buffer = BytesIO()
        group_name = group.nama if group else "Output"

        with ZipFile(zip_buffer, "w") as zf:
            for week in draft.weeks:
                if week.minggu in selected_weeks:
                    week_buffer = _generate_single_week_buffer(week, draft, group)
                    filename = f"RPP_Mingguan_{draft.metadata.kod_kursus or 'RPP'} - Minggu {week.minggu}.xlsx"
                    zf.writestr(filename, week_buffer.getvalue())

        zip_buffer.seek(0)
        result[group_name] = zip_buffer

    return result


def generate_combined_zip(draft: SessionDraft, selected_weeks: list[int]) -> BytesIO:
    """Generate a single ZIP with sub-folders per kumpulan."""
    combined_buffer = BytesIO()
    groups = draft.kumpulan_list if draft.kumpulan_list else [None]

    with ZipFile(combined_buffer, "w") as zf:
        for group in groups:
            group_name = group.nama if group else "Output"
            for week in draft.weeks:
                if week.minggu in selected_weeks:
                    week_buffer = _generate_single_week_buffer(week, draft, group)
                    filename = f"{group_name}/RPP_Mingguan_{draft.metadata.kod_kursus or 'RPP'} - Minggu {week.minggu}.xlsx"
                    zf.writestr(filename, week_buffer.getvalue())

    combined_buffer.seek(0)
    return combined_buffer
